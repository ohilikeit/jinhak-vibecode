#!/usr/bin/env python3
"""
jinhak-harness — common/utils/xlsx-write
템플릿 xlsx를 복사하고 row 2부터 데이터 행을 채워 새 xlsx를 만든다.

사용:
    python3 write_from_template.py <template.xlsx> <output.xlsx> <rows.json>

rows.json: 객체 배열. 키 = 템플릿 row 1의 헤더 이름.
출력 (stdout JSON): {"rows_written": <int>, "headers": [...], "output": "<path>"}
"""
from __future__ import annotations

import json
import shutil
import sys
from pathlib import Path


def fill(template: Path, output: Path, rows: list[dict]) -> dict:
    import openpyxl  # lazy

    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template, output)

    wb = openpyxl.load_workbook(output)
    ws = wb.active

    # row 1 헤더 읽기 (None 컬럼은 무시)
    headers: list[str] = []
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            break
        headers.append(str(cell.value))

    if not headers:
        raise SystemExit("템플릿 row 1에 헤더가 없습니다")

    # row 2부터 채우기
    for r_offset, row in enumerate(rows):
        if not isinstance(row, dict):
            raise SystemExit(f"row {r_offset}: 객체(dict)가 아닙니다")
        target_row = 2 + r_offset
        for c_idx, header in enumerate(headers, start=1):
            ws.cell(row=target_row, column=c_idx, value=row.get(header))

    wb.save(output)

    return {
        "rows_written": len(rows),
        "headers": headers,
        "output": str(output),
    }


def main(argv: list[str]) -> int:
    if len(argv) < 4:
        print(
            "usage: write_from_template.py <template.xlsx> <output.xlsx> <rows.json>",
            file=sys.stderr,
        )
        return 2

    template = Path(argv[1])
    output = Path(argv[2])
    rows_path = Path(argv[3])

    if not template.exists():
        print(f"template not found: {template}", file=sys.stderr)
        return 1
    if not rows_path.exists():
        print(f"rows.json not found: {rows_path}", file=sys.stderr)
        return 1

    rows = json.loads(rows_path.read_text(encoding="utf-8"))
    if not isinstance(rows, list):
        print("rows.json must be a list", file=sys.stderr)
        return 1

    result = fill(template, output, rows)
    json.dump(result, sys.stdout, ensure_ascii=False)
    sys.stdout.write("\n")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
