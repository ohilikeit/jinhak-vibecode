#!/usr/bin/env python3
"""
jinhak-harness — common/utils/xlsx-write inspector
xlsx 파일을 읽어 검증용 구조를 stdout JSON으로 출력한다.

사용:
    python3 inspect.py <file.xlsx>

출력:
    {
      "headers": ["공고제목", ...],
      "rows": [[...], ...],         # row 2부터
      "empty_cells": [[row, col, header], ...],
      "header_bold": [true, ...],
      "row_count": <int>
    }
"""
from __future__ import annotations

import json
import sys
from pathlib import Path


def inspect(xlsx: Path) -> dict:
    import openpyxl  # lazy

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active

    # 헤더 (None 만나면 종료)
    headers: list[str] = []
    for cell in ws[1]:
        if cell.value is None:
            break
        headers.append(str(cell.value))

    header_bold = [bool(c.font.bold) for c in ws[1][: len(headers)]]

    rows: list[list] = []
    empty_cells: list[tuple[int, int, str]] = []

    for row_idx in range(2, ws.max_row + 1):
        row_values = []
        any_value = False
        for col_idx in range(1, len(headers) + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            row_values.append(v)
            if v not in (None, ""):
                any_value = True
        if not any_value:
            continue
        rows.append(row_values)
        for col_idx, v in enumerate(row_values, start=1):
            if v in (None, ""):
                empty_cells.append([row_idx, col_idx, headers[col_idx - 1]])

    return {
        "headers": headers,
        "rows": rows,
        "empty_cells": empty_cells,
        "header_bold": header_bold,
        "row_count": len(rows),
    }


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("usage: inspect.py <file.xlsx>", file=sys.stderr)
        return 2
    xlsx = Path(argv[1])
    if not xlsx.exists():
        print(f"file not found: {xlsx}", file=sys.stderr)
        return 1
    data = inspect(xlsx)
    json.dump(data, sys.stdout, ensure_ascii=False)
    sys.stdout.write("\n")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
